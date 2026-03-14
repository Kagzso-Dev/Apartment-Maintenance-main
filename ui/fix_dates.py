import openpyxl
from datetime import datetime
import os

def format_date_to_string(value):
    """Convert any date format to DD-MM-YYYY string"""
    if value is None or value == '':
        return ''
    
    # If it's already a string in the correct format, return as is
    if isinstance(value, str):
        if value.strip() and '-' in value:
            # Check if it's already in DD-MM-YYYY format
            try:
                parts = value.split('-')
                if len(parts) == 3 and len(parts[2]) == 4:
                    return value.strip()
            except:
                pass
    
    # If it's a datetime object, convert to DD-MM-YYYY
    if isinstance(value, datetime):
        return value.strftime('%d-%m-%Y')
    
    # Try parsing as datetime string
    try:
        dt = datetime.strptime(str(value)[:10], '%Y-%m-%d')
        return dt.strftime('%d-%m-%Y')
    except:
        pass
    
    return str(value)

def fix_excel_dates(filepath):
    """Fix all date columns in Excel to be text strings"""
    print(f"Opening file: {filepath}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Find the PaymentDate column (usually column J)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    payment_date_col = None
    
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == 'PaymentDate':
            payment_date_col = idx
            break
    
    if not payment_date_col:
        print("PaymentDate column not found!")
        return
    
    print(f"PaymentDate column found at index {payment_date_col}")
    
    # Process all rows
    fixed_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        cell = row[payment_date_col - 1]
        original_value = cell.value
        
        if original_value:
            new_value = format_date_to_string(original_value)
            if new_value != original_value:
                cell.value = new_value
                cell.number_format = '@'  # Set as text format
                fixed_count += 1
                print(f"Row {row_idx}: {original_value} -> {new_value}")
    
    # Save with backup
    backup_path = filepath.replace('.xlsx', '_backup.xlsx')
    if not os.path.exists(backup_path):
        print(f"Creating backup: {backup_path}")
        wb.save(backup_path)
    
    print(f"Saving changes to: {filepath}")
    wb.save(filepath)
    print(f"✓ Fixed {fixed_count} date cells")
    print(f"✓ Backup saved as: {backup_path}")

if __name__ == '__main__':
    file_path = './data/Monthly_Transactions.xlsx'
    fix_excel_dates(file_path)
